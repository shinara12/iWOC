from fpdf import FPDF
import pandas as pd
import openpyxl
from datetime import datetime
from datetime import datetime
from datetime import datetime
from barcode import EAN13,Code39
from barcode.writer import ImageWriter
from fpdf import FPDF
from pyzbar.pyzbar import decode
from PIL import Image
import PyPDF2
import os
from datetime import datetime,date,timedelta
import datetime
from .bkrcc import *
def rotate_func(kk):
    pdf_in = open(r"%s\pgi11.pdf"%kk, 'rb')
    pdf_reader = PyPDF2.PdfFileReader(pdf_in)
    pdf_writer = PyPDF2.PdfFileWriter()

    for pagenum in range(pdf_reader.numPages):
        page = pdf_reader.getPage(pagenum)
        if pagenum % 2:
            page.rotateClockwise(180)
        pdf_writer.addPage(page)

    pdf_out = open(r"%s\pgi.pdf"%kk, 'wb')
    pdf_writer.write(pdf_out)
    pdf_out.close()
    pdf_in.close() 
    
    import os
    os.remove(r"%s\pgi11.pdf"%kk)

def PGI(request,files1):
    move_files = []
    move_files.append(files1)
    p = 1
    cppp = 1
    def Folder_creation(cppp):
        # ------------------Capture date----------------
        cppp=str(cppp)
        today=date.today()
        d1 = today.strftime("%d-%b-%Y")
        # print(d1,'ddddddddddddddddddddddddddddddddd')
        d2=d1[0:2]
        d3=d1[3:6]
        d1=d2+d3+cppp
        # ------------------Capture time----------------
        now = datetime.now()        
        start_time = now.strftime("%H:%M:%S")
        s1=start_time[0:2]
        s2=start_time[3:5]
        s3=start_time[6:8]
        start_time=s1+s2+s3

        # ---------------create first date folder------------
        # import os
        # directory = d
        # print('Helloooooooooooooooooooooooooooooooo')
        directory=d1
        parent_dir = "E:\IWOC\OUTPUT\Bank Simpanan Nasional\CYCLE\\"
        directory = d1+"_"+start_time 
        path1 = os.path.join(parent_dir, directory) # date folder path
        os.mkdir(path1)
        # print("Directory '% s' created" % directory)

        directory = 'pgi'
        parent_dir = path1
        path2 = os.path.join(path1, directory) # bkrcc folder path
        os.mkdir(path2)
        # print("Directory '% s' created" % directory)

        directory = 'PRINTING'
        parent_dir = path2
        path3 = os.path.join(path2, directory) # DIGITAL folder path
        os.mkdir(path3)
        # print("Directory '% s' created" % directory)

        # directory = 'PRINTING'
        # parent_dir = path2
        # path4 = os.path.join(path2, directory) # PRINTING folder path
        # os.mkdir(path4)
        # print("Directory '% s' created" % directory)

        return path3
    
    path11=r'E:\IWOC\SOURCE\\'+ files1
    data=pd.read_excel(path11)
    number_of_rows = len(data)
    data['br_code']=data['ILOM_BRCH_CODE']
    data['b_date']=data['BDATE']
    data['name']=data['NAME']
    data['add1']=data['ADDR1']
    data['add2']=data['ADDR2']
    data['add3']=data['ADDR3']
    data['add4']=data['ADDR4']
    data['postcode']=data['POSTCODE']
    data['city_cd']=data['CITY_CD']
    data['state']=data['STATE']
    data['acc_no']=data['ILOM_OLD_SEQUENCE']
    data['sub_code_desc']=data['ILPC_SUBCODE_DESC']
    data['duedt_pgi']=data['DUEDT_PGI']
    data['total_due']=data['TOTAL_DUE']
    data['amtdue_pgi'] =data['AMTDUE_PGI']
    data['memo_fees']=data['MEMO_FEES']
    data['duedt_pgi'] = data['duedt_pgi'].dt.strftime('%d/%m/%Y')
    
    l1 = data['br_code'].to_list()
    br_f = []
    for i in l1:
        i = str(i)
        if len(i)  == 3:
            br_f.append('00' + i)
        else:
            br_f.append('0' + i)


    amt_due_pgi = []
    amt_due_pgi_final = []
    amt_due_pgi = data['amtdue_pgi'].to_list()
    for i in amt_due_pgi:
        i = str(i)
        k = i.split(".")
        k0 = k[0]
        k1 = k[-1]
        if len(k1) == 1:
            k1 = k[-1] + '0'
        k2 = k0 + '.' + k1
        amt_due_pgi_final.append(k2)
    
    data['memo_fees'] = data['memo_fees']
    memo_fees = []
    memo_fees_final = []
    memo_fees = data['memo_fees'].to_list()
    for i in memo_fees:
        i = str(i)
        k = i.split(".")
        k0 = k[0]
        k1 = k[-1]
        if len(k1) == 1:
            k1 = k[-1] + '0'
        k2 = k0 + '.' + k1
        memo_fees_final.append(k2)
    
    data['total_due'] = data['total_due']
    total_due = []
    total_due_final = []
    total_due = data['total_due'].to_list()
    for i in total_due:
        i = str(i)
        k = i.split(".")
        k0 = k[0]
        k1 = k[-1]
        if len(k1) == 1:
            k1 = k[-1] + '0'
        k2 = k0 + '.' + k1
        total_due_final.append(k2)

    time1 = datetime.now() 
    start_time=time1.strftime("%H:%M:%S")
    date1 = time1.strftime("%d/%b/%Y")

    file=open(r"E:\\IWOC\\LOG_FILES\\pgi.txt","a")
    file.write("INTERCITY  MPC (M) Sdn. Bhd.\n")
    file.write("FIRST REMINDER\n")
    file.write("=============================================================\n")
    file.write("Program ID           : FIRST REMINDER.PDS\n")
    file.write("Paper Type           :\n")
   
    def barcode(acc_no,count):
        now = datetime.now() 
        date22 = now.strftime("%m%d%y")
        barcode_no = '0906' +  str(date22) +  str(acc_no) 
        # print(barcode_no,'barcode_nobarcode_nobarcode_nobarcode_no')
        # print(acc_no,'accccccccccccccccccccccnooooooooooooooooo')

        number = barcode_no
        my_code = Code39(number, writer=ImageWriter())
        new_code  = 'b' + str(count)
        my_code.save(r"E:\IWOC\EXTRA\barcode_fo\%s" %(new_code))


        img = Image.open(r"E:\IWOC\EXTRA\barcode_fo\%s.png" %(new_code))
        imgCropped = img.crop(box=(4,10,1080,200))
        imgCropped.save(r"E:\IWOC\EXTRA\barcode_fo\%s.png" %(new_code))

    # print(data['amtdue_pgi'],'------------------------------------')
    count  = 1
    c = 1
    c1 = 1
    pdf = FPDF()
    for i in range(0,number_of_rows):
        # print(i,data['acc_no'][i],'0000000000000000000000000000')
        barcode(data['acc_no'][i],count)
        pdf.add_page()
        pdf.set_auto_page_break(auto=False)
        pdf.set_line_width(6)
        pdf.set_draw_color(r=0, g=0, b=0)
        pdf.line(x1=5, y1=4, x2=205, y2=4)         #1st horizontal line
        pdf.line(x1=5, y1=292, x2=205, y2=292)     #2nd horizontal line 
        pdf.line(5,6,5,290)                         # 1st vertical line
        pdf.line(205,6,205,290)                     # 2nd vertical line    

        pdf.set_font('Times','B',8.5)
        pdf.cell(45)
        pdf.cell(10,10,txt = 'PERATURAN-PERATURAN PEMBAYARAN PINJAMAN/PEMBIAYAAN')
        pdf.ln(1)
        pdf.cell(66)
        pdf.cell(10,14,txt = 'BANK SIMPANAN NASIONAL (BSN)')

        pdf.ln(2)
        pdf.set_font('Times','',8.5)
        pdf.cell(10,30,txt ='Bayaran PINJAMAN / PEMBIAYAAN boleh dibuat secara tunai di mana-mana ATM/CDM/Cawangan BSN, pindahan GIRO / GIRO-I Perbankan internet') 
        pdf.ln(1)
        pdf.cell(10,36, txt = '(www.mybsn.com.my) dan Inter Bank Giro (IBG) dari bank-bank lain yang mengambil bahagian.')


        pdf.ln(4)
        pdf.set_font('Times','',8.5)
        pdf.cell(10,44,txt ='Bayaran PINJAMAN / PEMBIAYAAN juga boleh dibuat menerusi arahan Tetap akaun simpanan GIRO/GIRO-I.Bagi membolehkan kami mendebit Akaun') 
        pdf.ln(1)
        pdf.cell(10,50, txt = 'Simpanan BSN tuan/puan, sila pastikan amaun tuan/puan mempunyai baki minima yang ditetapkan oleh BSN dan amaun yang perlu dijelaskan.')


        pdf.ln(4)
        pdf.set_font('Times','',8.5)
        pdf.cell(10,58,txt ='Sekiranya pembayaran melalui cek, draf bank atau arahan juruwang hendaklah dibuat atas nama') 
        pdf.cell(106)
        pdf.set_font('Times','B',8.5)
        pdf.cell(10,58,txt ='BANK SIMPANAN NASIONAL') 
        pdf.cell(33)
        pdf.set_font('Times','',8.5)
        pdf.cell(10,58,txt ='dan dipalang') 
        pdf.cell(5.8)
        pdf.set_font('Times','B',8.5)
        pdf.cell(10,58,txt ='"AKAUN') 
        pdf.ln(1)
        pdf.set_font('Times','B',8.5)
        pdf.cell(10,64, txt = 'PENERIMA SAHAJA"')
        pdf.set_font('Times','',8.4)
        pdf.cell(20)
        pdf.cell(10,64, txt = '. Catatkan nama dan nombor akaun Pinjaman / Pembiayaan tuan/puan di belakang cek berkenaan. Bayaran hendaklah sampai kepada')
        pdf.ln(1)
        pdf.cell(10,70, txt = 'kami sebelum tarikh akhir yang ditetapkan.')


        pdf.ln(4)
        pdf.set_font('Times','',9)
        pdf.cell(10,76,txt ='Denda lewat bayar / caj perkhidmatan akan dikenakan jika bayaran yang dibuat oleh tuan / puan diterima selepas tarikh akhir yang ditetapkan. Jika') 
        pdf.ln(1)
        pdf.cell(10,82, txt = 'tuan/puan telah menjelaskan amaun berkenaan, sila abaikan pemberitahuan ini.')

        pdf.ln(4)
        pdf.set_font('Times','',9)
        pdf.cell(10,88,txt ='Jika tuan / puan memerlukan penjelasan selanjutnya berhubung butir / butir yang dinyatakan di hadapan, sila hubungi kami di talian 03-2028 3222') 
        pdf.ln(1)
        pdf.cell(10,94, txt = 'atau cawangan di mana tuan / puan membuat permohonan pinjaman / pembiayaan tersebut di talian-talian berikut:-')

        pdf.ln(1)
        pdf.cell(10,104,txt = 'Johor')
        pdf.cell(15)
        pdf.cell(10,104,txt = 'Kedah')
        pdf.cell(15)
        pdf.cell(10,104,txt = 'Kelantan')
        pdf.cell(15)
        pdf.cell(10,104,txt = 'N. Sembilan')
        pdf.cell(15)
        pdf.cell(10,104,txt = 'Melaka')
        pdf.cell(15)
        pdf.cell(10,104,txt = 'Pahang')

        pdf.ln(1)
        pdf.cell(10,108,txt = '07-208 3555')
        pdf.cell(15)
        pdf.cell(10,108,txt = '04-7740 444')
        pdf.cell(15)
        pdf.cell(10,108,txt = '09-745 7070')
        pdf.cell(15)
        pdf.cell(10,108,txt = '06-768 6500')
        pdf.cell(15)
        pdf.cell(10,108,txt = '06-289 5800')
        pdf.cell(15)
        pdf.cell(10,108,txt = '09-565 0565')


        pdf.ln(4)
        pdf.cell(10,110,txt = 'Perak')
        pdf.cell(15)
        pdf.cell(10,110,txt = 'P.Pinanag')
        pdf.cell(15)
        pdf.cell(10,110,txt = 'Sabah')
        pdf.cell(15)
        pdf.cell(10,110,txt = 'Saeawak')
        pdf.cell(15)
        pdf.cell(10,110,txt = 'Terengganu')
        pdf.cell(15)
        pdf.cell(10,110,txt = 'Selangor')

        pdf.ln(1)
        pdf.cell(10,114,txt = '05-245 2222')
        pdf.cell(15)
        pdf.cell(10,114,txt = '04-222 6400')
        pdf.cell(15)
        pdf.cell(10,114,txt = '088-355 600')
        pdf.cell(15)
        pdf.cell(10,114,txt = '082-227 888')
        pdf.cell(15)
        pdf.cell(10,114,txt = '09-6200 400')
        pdf.cell(15)
        pdf.cell(10,114,txt = '03-5543 3000')

        pdf.ln(1)
        pdf.set_line_width(1)
        pdf.line(10,105,200,105)

        pdf.ln(1)
        pdf.set_font('Courier','B',8)
        pdf.cell(90)
        pdf.cell(10,132,txt = 'SULIT')

        pdf.ln(1)
        pdf.cell(10,140, txt = 'JABATAN KHIDMAT PENGURUSAN &')
        pdf.cell(140)
        pdf.cell(10,140,txt = 'TANPA PRASANGKA')

        pdf.ln(1)
        pdf.cell(10,146,txt = 'PEMULIHAN HUTANG')

        pdf.ln(1)
        pdf.cell(10,152,txt = 'BANK SIMPANAN NASIONAL')


        pdf.ln(1)
        pdf.cell(2)
        pdf.cell(10)
        pdf.cell(10,156,txt ="(")
        pdf.cell(-8)
        pdf.cell(12,156,txt = str(br_f[i]))
        pdf.cell(-3)
        pdf.cell(20,156,txt = ")")


        pdf.ln(1)
        pdf.cell(160)
        a = data['b_date'][i]
        a = data['b_date'].apply(lambda x: x.strftime('%d/%m/%Y').split('-'))
        a = "".join(a[0])
        pdf.cell(10,156,txt = str(a))

        h = 170
        if pd.notna(data['name'][i]):
            pdf.ln(1)
            pdf.cell(60)
            pdf.cell(10,h,txt = str(data['name'][i]))

        if pd.notna(data['add1'][i]):
            h = h + 6
            pdf.ln(1)
            pdf.cell(60)
            pdf.cell(10,h,txt = str(data['add1'][i]))

        if pd.notna(data['add2'][i]):
            h = h + 6
            pdf.ln(1)
            pdf.cell(60)
            pdf.cell(10,h,txt = str(data['add2'][i]))

        if pd.notna(data['add3'][i]):
            h = h + 6
            pdf.ln(1)
            pdf.cell(60)
            pdf.cell(10,h,txt = str(data['add3'][i]))

        if pd.notna(data['add4'][i]):
            h = h + 6
            pdf.ln(1)
            pdf.cell(60)
            pdf.cell(10,h,txt = str(data['add4'][i]))

        if pd.notna(data['postcode'][i]):
            h = h + 6
            pdf.ln(1)
            pdf.cell(60)
            pdf.cell(10,h,txt = str(data['postcode'][i]))
            pdf.cell(1)
            pdf.cell(10,h,txt =str(data['city_cd'][i]))

        if pd.notna(data['state'][i]):
            h = h + 6
            pdf.ln(1)
            pdf.cell(60)
            pdf.cell(10,h,txt = str(data['state'][i]))

        pdf.ln(1)
        pdf.cell(10)
        pdf.cell(10,220, txt = 'Account Number: ')


        pdf.cell(16)
        pdf.cell(10,220, txt = str(data['acc_no'][i]))


        pdf.cell(50)
        pdf.cell(10,220, txt = str(data['sub_code_desc'][i]))
        pdf.ln(1)
        # pdf.cell(100)
        # pdf.cell(10,226, txt = '3.95)')


        pdf.ln(1)
        pdf.cell(10)
        pdf.cell(10,240,txt = 'PEMBERITAHUAN PEMBAYARAN FAEDAH BULANAN.')

        pdf.ln(1)
        pdf.cell(10)
        pdf.cell(10,254,txt = 'BUTIR-BUTIR PEMBIAYAAN/PINJAMAN TUAN/PUAN SEHINGGA TARIKH PEMBERITAHUAN INI')


        pdf.ln(1)
        pdf.cell(10)
        pdf.cell(10,260,txt = 'ADALAH :')


        pdf.ln(1)
        pdf.cell(10)
        pdf.cell(10,270,txt = 'FAEDAH BULANAN')

        pdf.cell(30)
        pdf.cell(10,270,txt = str(data['duedt_pgi'][i]))
        pdf.cell(40)
        pdf.cell(2)
        pdf.cell(10,270,txt = ': RM')
        # pdf.cell(10,270,txt = str(data['amtdue_pgi'][i]))
        pdf.cell(10,270,txt = str(amt_due_pgi_final[i]))
        # print(str(data['amtdue_pgi'][i]),'****************************************((()(')

        pdf.ln(1)
        pdf.cell(10)
        pdf.cell(10,276,txt = 'TUNGGAKAN')

        pdf.cell(80)
        pdf.cell(2)
        pdf.cell(10,276,txt = ': RM')
        pdf.cell(10,276,txt = str(memo_fees_final[i]))

        pdf.ln(1)
        pdf.cell(10)
        pdf.cell(10,282,txt = 'AMAUN YANG PERLU DIBAYAR')

        pdf.cell(80)
        pdf.cell(2)
        pdf.cell(10,282,txt = ': RM')
        pdf.cell(10,282,txt = str(total_due_final[i]))

        pdf.ln(1)
        pdf.cell(10)
        pdf.cell(10,300,txt = 'SILA JELASKAN AMAUN YANG PERLU DIBAYAR TERSEBUT DALAM TEMPOH 14 HARI DARI')

        pdf.ln(1)
        pdf.cell(10)
        pdf.cell(10,306,txt = 'TARIKH NOTIS INI. SILA HUBUNGI KAMI DI TALIAN-TALIAN YANG DINYATAKAN PADA MUKA')

        pdf.ln(1)
        pdf.cell(10)
        pdf.cell(10,312,txt = 'SURAT DI ATAS JIKA TUAN/PUAN MEMERLUKAN PENJELASAN SELANJUTNYA.')


        pdf.ln(1)
        pdf.cell(10)
        pdf.cell(10,324 ,txt = 'AGENSI KAUNSELING DAN PENGURUSAN KREDIT TELAH DITUBUHKAN OLEH BANK NEGARA')

        pdf.ln(1)
        pdf.cell(10)
        pdf.cell(10,330 ,txt = 'MALAYSIA UNTUK MENYEDIAKAN PERKHIDMATAN PENGURUSAN KEWANGAN, KAUNSELING KREDIT,')

        pdf.ln(1)
        pdf.cell(10)
        pdf.cell(10,336 ,txt = 'PENDIDIKAN KEWANGAN DAN PENSTRUKTURAN SEMULA PINJAMAN SECARA PERCUMA KEPADA')

        pdf.ln(1)
        pdf.cell(10)
        pdf.cell(10,342 ,txt = 'INDIVIDU. UNTUK MEMBUAT PERTANYAAN, SILA HUBUNGI TALIAN 1-800-88-2575')


        pdf.ln(1)
        pdf.cell(10)
        pdf.cell(10,352 ,txt = 'NOTA: SEKIRANYA TUAN/PUAN DIBAWAH PROGRAM PENANGGUHAN BAYARAN (MORATORIUM).')

        pdf.ln(1)
        pdf.cell(20)
        pdf.cell(10,356 ,txt = 'SILA IKUTI TARIKH PEMBAYARAN SEPERTI YANG TELAH DITETAPKAN.')


        pdf.ln(1)
        pdf.cell(10)
        pdf.set_font('Courier','B',9)
        pdf.cell(10,375 ,txt = 'Ingin bayar bil dan semak baki akaun simpanan anda secara online? Daftarlah')


        pdf.ln(1)
        pdf.cell(10)
        pdf.cell(10,381 ,txt = 'Perbankan Internet MyBSN sekarang! Kemudahan perbankan hanya untuk anda dari BSN.')


        pdf.ln(1)
        pdf.cell(10)
        pdf.cell(10,387 ,txt = 'Untuk maklumat lanjut, layari www.mybsn.com.my atau 1300 88 1900.')


        pdf.ln(1)
        pdf.cell(140)
        pdf.set_font('Courier','B',6)
        pdf.cell(10,410 ,txt = 'PGI-01032022_030322.csv-00000' + str(count))

        pdf.add_page()
        pdf.image(r'E:\IWOC\APP SCRIPT\PGI\CYCLE\PGI\DATA\BSN_Reminder-01.jpg',-10,-10,230)

        pdf.set_font('Courier','',7)
        pdf.cell(10,280,txt = '00' + str(c))
        c += 1
        pdf.ln(1)
        pdf.image(r'E:\IWOC\APP SCRIPT\PGI\CYCLE\PGI\DATA\BSN_Reminder-02.jpg',20,146,80)
        pdf.image(r'E:\IWOC\APP SCRIPT\PGI\CYCLE\PGI\DATA\BSN_Reminder-02 - Copy.jpg', 145,146,45)

        pdf.ln(1)
        # pdf.image('E:\IWOC\SOURCE\sampleqcode.PNG',40,210,60)
        new_code  = 'b' + str(count) 
        pdf.image(r'E:\IWOC\EXTRA\barcode_fo\%s.png'%(new_code),40,210,40)
        

        pdf.ln(1)
        pdf.cell(30)
        pdf.set_font('Courier','',7)
        pdf.cell(10,414,txt = '00000' +str(c1) +  '-P06-01')
        c1 += 1
        h = 420
        if pd.notna(data['name'][i]):
            pdf.ln(1)
            pdf.set_font('Courier','B',10)
            pdf.cell(30)
            pdf.cell(10,h,txt = str(data['name'][i]))

        if pd.notna(data['add1'][i]):
            h = h + 6
            pdf.ln(1)
            pdf.cell(30)
            pdf.cell(10,h,txt = str(data['add1'][i]))
        
        if pd.notna(data['add2'][i]):
            h = h + 6
            pdf.ln(1)
            pdf.cell(30)
            pdf.cell(10,h,txt = str(data['add2'][i]))

        if pd.notna(data['add3'][i]):
            h = h + 6
            pdf.ln(1)
            pdf.cell(30)
            pdf.cell(10,h,txt = str(data['add3'][i]))

        if pd.notna(data['add4'][i]):
            h = h + 6
            pdf.ln(1)
            pdf.cell(30)
            pdf.cell(10,h,txt = str(data['add4'][i]))
        
        if pd.notna(data['postcode'][i]):
            h = h + 6
            pdf.ln(1)
            pdf.cell(30)
            pdf.cell(10,h,txt = str(data['postcode'][i]) )
            pdf.cell(2)
            pdf.cell(10,h,str(data['city_cd'][i]))

        if pd.notna(data['state'][i]):
            h = h + 6
            pdf.ln(1)
            pdf.cell(30)
            pdf.cell(10,h,txt = str(data['state'][i]))
        count = int(count)
        count += 1


    kk=Folder_creation(cppp)
    bk = kk
   
    # print(kk,'kkkkkkkkkkkkkkkkkkkkkkkkkkkkk')
        
    pdf.output(r"%s\pgi11.pdf"%kk)
    if i == number_of_rows-1:
        rotate_func(bk)
    lesson =(r"%s\pgi11.pdf"%kk)
    file_name = os.path.split(lesson)
    ffname = file_name[-1]
    # print(ffname,'saaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaam')
    cppp += 1
    
    file.write("File Name (PDF)      :      ")
    file.write(str(ffname))
    file.write("\n")
    file.write("Data Name (TXT)      :      ")
    file.write(str(files1))
    file.write("\n")
    file.write("Total Accounts       :        ")
    file.write(str(number_of_rows))
    file.write("\n")
    file.write("Total Pages          :        ")
    file.write(str(number_of_rows))
    file.write("\n")
    file.write("Total Impression     :        ")
    file.write(str(number_of_rows*2))
    file.write("\n")

    file.write("First Record         : ")
    file.write(str(data['acc_no'][0]))
    file.write("\n")
    file.write("Last Record          : ")
    last_record=data['acc_no'].iat[-1]
    file.write(str(last_record))
    file.write("\n")
    file.write("=============================================================\n")
    file.write("CUSTOMER NAME        : BSN                                   \n")
    file.write("=============================================================\n")
    file.write("Date & Time processed: ")


    file.write(str(date1))
    file.write("    ")
    file.write(str(start_time))
    file.write("    " )
    
    # pdf.output(r'E:\IWOC\APP SCRIPT\PGI\CYCLE\PGI\DATA\samplepgi.pdf')
    time1 = datetime.now()
    end_time=time1.strftime("%H:%M:%S")
    file.write("    ")
    file.write(str(end_time) )
    file.write('\n\n\n\n')
    file.close()
   


    path20 = r'E:\IWOC\SOURCE' 
    target = r'E:\IWOC\TEMP' 

    for i in move_files:

        src_path = os.path.join(path20, i)
        dst_path = os.path.join(target, i)
        os.rename(src_path, dst_path)
    print('Files Moved................')







