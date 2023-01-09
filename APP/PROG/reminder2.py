import PyPDF2
import pandas as pd
import openpyxl
import datetime
from fpdf import FPDF
from datetime import datetime
from datetime import datetime
from barcode import EAN13,Code39
from barcode.writer import ImageWriter
from fpdf import FPDF
from pyzbar.pyzbar import decode
from PIL import Image
from PIL import Image
import os
import datetime
from .bkrcc import *
from datetime import datetime,timedelta
from datetime import date
from .bkrcc import *
def rotate_func(kk):
    pdf_in = open(r"%s\rem22.pdf"%kk, 'rb')
    pdf_reader = PyPDF2.PdfFileReader(pdf_in)
    pdf_writer = PyPDF2.PdfFileWriter()

    for pagenum in range(pdf_reader.numPages):
        page = pdf_reader.getPage(pagenum)
        if pagenum % 2:
            page.rotateClockwise(180)
        pdf_writer.addPage(page)

    pdf_out = open(r"%s\rem2.pdf"%kk, 'wb')
    pdf_writer.write(pdf_out)
    pdf_out.close()
    pdf_in.close() 
    
    import os
    os.remove(r"%s\rem22.pdf"%kk)

def REMINDER2(request,files1):
    move_files = []
    move_files.append(files1)
    p = 1
    cppp = 1
    def Folder_creation(cppp):
        # ------------------Capture date----------------
        cppp=str(cppp)
        today=date.today()
        d1 = today.strftime("%d-%b-%Y")
        # print(d1,'ddddddddddddddddddddddddddddddddd')
        d2=d1[0:2]
        d3=d1[3:6]
        d1=d2+d3+cppp
        # ------------------Capture time----------------
        now = datetime.now()        
        start_time = now.strftime("%H:%M:%S")
        s1=start_time[0:2]
        s2=start_time[3:5]
        s3=start_time[6:8]
        start_time=s1+s2+s3

        # ---------------create first date folder------------
        # import os
        # directory = d
        # print('Helloooooooooooooooooooooooooooooooo')
        directory=d1
        parent_dir = "E:\IWOC\OUTPUT\Bank Simpanan Nasional\CYCLE\\"
        directory = d1+"_"+start_time 
        path1 = os.path.join(parent_dir, directory) # date folder path
        os.mkdir(path1)
        # print("Directory '% s' created" % directory)

        directory = 'BsnReminder2'
        parent_dir = path1
        path2 = os.path.join(path1, directory) # bkrcc folder path
        os.mkdir(path2)
        # print("Directory '% s' created" % directory)

        directory = 'PRINTING'
        parent_dir = path2
        path3 = os.path.join(path2, directory) # DIGITAL folder path
        os.mkdir(path3)
        # print("Directory '% s' created" % directory)

        # directory = 'PRINTING'
        # parent_dir = path2
        # path4 = os.path.join(path2, directory) # PRINTING folder path
        # os.mkdir(path4)
        # print("Directory '% s' created" % directory)

        return path3

    path11 = r'E:\IWOC\SOURCE\\'+ files1
    data=pd.read_excel(path11)
    # print(data.columns)
    df=pd.DataFrame()
    df['br_code']=data['ILOM_BRCH_CODE']
    df['CUSTOMER_NAME']=data['CUSTOMER_NAME']
    df['ADDR1']=data['ADDR1']
    df['ADDR2']=data['ADDR2']
    df['ADDR3']=data['ADDR3']
    df['ADDR4']=data['ADDR4']
    df['POSTCODE']=data['POSTCODE']
    df['CITY_CD']=data['CITY_CD']
    df['STATE']=data['STATE']
    df['ILOM_OLD_SEQUENCE']=data['ILOM_OLD_SEQUENCE']
    df['ICBS_PROD_SUBDESC']=data['ICBS_PROD_SUBDESC']
    df['TOTAL_DUE']=data['TOTAL_DUE']
    kk1=data['SYSDATE_GENERATE'].to_list()
    kk2=data['POSTCODE'].to_list()

    date0=[]
    date2 = []
    for i in kk1:
        date0.append(i.strftime("%d/%m/%Y"))
        date2.append(i.strftime("%d%m%y"))
    df['SYSDATE_GENERATE']=date0
    # print(df['SYSDATE_GENERATE'],'dateeeee')
    postcode=[]
    for i in kk2:
        # ----------1-10-----------
        if i>=1000 and i<=2999:
            postcode.append('001')
        elif i>=5000 and i<=6999:
            postcode.append('002')
        elif i>=7000 and i<=7999:
            postcode.append('003')
        elif i>=8000 and i<=8099:
            postcode.append('004')
        elif (i>=8100 and i<=8899) or (i>=9100 and i<=9599) or (i>=9700 and i<=9899) or i==34950:
            postcode.append('005')
        elif i>=9000 and i<=9099 or (i>=9600 and i<=9699):
            postcode.append('006')
        elif i>=10000 and i<=10999 or (i>=11000 and i<=1150) or i==11800 or (i>=13100 and i<=13310) or i==13050 or (i>=14100 and i<=14390):
            postcode.append('007')
        elif i==11600 or i==11700:
            postcode.append('008')
        elif i>=11900 and i<=11999:
            postcode.append('009')
        elif i>=12000 and i<=12300 or (i>=12700 and i<=12990) or (i>=13000 and i<=13020) or i==13400 or i==13800:
            postcode.append('010')
        
        # -----------11-20---------------------
        elif i == 13500 or i == 13600 or i == 13700:
            postcode.append('011')
        elif i>=14000 and i<= 14020 or i == 14400:
            postcode.append('012')
        elif i >= 30000 and i <= 30999:
            postcode.append('013')
        elif i == 31450 or i == 31500 or i == 31650:
            postcode.append('014')
        elif (i >= 31000 and i <= 31100) or i == 31200 or i == 31300 or i == 31350 or (i >= 31550 and i <=31610) or (i >=31700 and i <=32900) or (i >= 33100 and i <= 33500) or (i >= 34100 and i <= 34500) or i == 34800 or i == 34900 or (i >= 35000 and i <= 35999) or (i >= 39000 and i <= 39999):
            postcode.append('015')
        elif i == 31400 or i == 31150 or i ==  31250:
            postcode.append('016')
        elif  (i >= 33000 and i <= 33099) or i == 33600 or i == 33700:
            postcode.append('017') 
        elif (i >= 34000 and i <= 34099) or i == 34600 or i == 34650  or i == 34700 or i == 34750 or i == 34850:
            postcode.append('018')
        elif (i >= 36000 and i <= 36999):
            postcode.append('019')        
        elif i == 40150:
            postcode.append('020')

        # -----------21-30---------------------
        elif i == 40400 or (i>= 40460 and i <= 40690) or i == 42450:
            postcode.append('021')
        elif i == 40000 or i == 40100 or i == 40200 or i == 40300 or i == 40450:
            postcode.append('022')
        elif (i >= 40700 and i <= 40990):
            postcode.append('023')
        elif i == 41000  or i == 41100 or i == 41200 or i == 41250 or (i >= 41500 and i <= 41990):
            postcode.append('024')
        elif i == 40170 or i == 41050 or i == 41150 or i == 41300 or i == 41400 or i == 42100:
            postcode.append('025')
        elif (i >= 42000 and i <=42009) or i == 42920:
            postcode.append('026')
        elif i == 42600 or i == 42700 or (i >=45000 and i <= 45030) or i == 45100:
            postcode.append('027')
        elif i == 42500 or i == 42800 or i == 45800 or (i >= 44000 and  i <= 44300) or i == 48100 or i == 42940 or i == 42960 or i == 45200 or i == 45300 or i == 45400 or i == 45500 or i == 45600 or i == 45700:
            postcode.append('028')
        elif i == 43900 or i == 43950 or i == 64000:
            postcode.append('029')
        elif (i >= 43000 and i <= 43009) or (i >= 43600 and i <= 43650) or (i >= 62000 and i <= 62306):
            postcode.append('030')

        # -----------31-40---------------------
        elif i == 43100 or (i >= 43200 and i <= 43207):
            postcode.append('031')
        elif (i >= 43300 and i <= 43499):
            postcode.append('032')
        elif i == 43500 or i == 43700:
            postcode.append('033')
        elif (i >= 46000 and i <= 46999):
            postcode.append('034')
        elif i == 40160 or i == 47000 or i == 47810 or i == 48050:
            postcode.append('035')
        elif (i >= 47100  and i <= 47190) or i == 42610:
            postcode.append('036')
        elif i == 47200 or (i >= 47300 and i <= 47308):
            postcode.append('037')
        elif i == 47400 or i == 47410 or i == 47800 or i == 47820 or i == 47830:
            postcode.append('038')
        elif (i >= 47500 and i <= 47699):
            postcode.append('039')
        elif i == 48000 or i == 48010 or i == 48020 or (i >= 48200 and i <= 48399):
            postcode.append('40')


        # -----------41-50---------------------
        elif (i >= 62400 and i <= 62688) or i == 62692:
            postcode.append('041')
        elif (i >=42200 and i <= 42300):
            postcode.append('042')
        elif (i >= 43800 and i <= 43809) or (i >= 63000 and i <= 63300):
            postcode.append('043')
        elif (i >= 50000 and i <= 50400):
            postcode.append('044')
        elif (i >= 50500 and i <= 50699):
            postcode.append('045')
        elif (i >= 50450 and i <= 50490):
            postcode.append('046')
        elif (i >= 50700 and i <= 50990):
            postcode.append('047')
        elif (i >= 51000 and i <= 51200) or ( i >= 51700 and i <= 51990):
            postcode.append('048')
        elif i == 52000 or ( i >= 52100 and i <= 52200):
            postcode.append('049')
        elif (i >= 53000 and i <= 53300) or (i >= 53700 and i <= 53990):
            postcode.append('050')

        
        # -----------51-60---------------------
        elif (i >= 54000 and i <= 54200):
            postcode.append('051')
        elif (i >= 55000 and i <= 55300) or (i >= 55700 and i <= 55990):
            postcode.append('052')
        elif (i >= 56000 and i <= 56100):
            postcode.append('053')
        elif (i >= 58000 and i <= 58200) or (i >= 58700 and i <= 58990):
            postcode.append('054')
        elif (i >= 590000 and i <= 59200)  or (i >= 59700 and i <= 59900):
            postcode.append('055')
        elif (i >= 57000 and i <= 57100) or (i >= 57700 and i <= 57990):
            postcode.append('056')
        elif i == 60000:
            postcode.append('057')    
        elif i == 68000:
            postcode.append('058')
        elif i == 68100:
            postcode.append('059')
        elif i == 70000 or (i >= 70100 and i <= 70200) or (i >= 70500  and i <= 70990) or i == 71450 or i == 71770:
            postcode.append('060')


        # -----------61-70---------------------                 
        elif i == 70300 or(i >= 70400 and i <= 70450):
            postcode.append('061')
        elif (i >=71000 and i <=71009) or i == 71010 or i == 72100 or i == 71760 or (i >= 71800 and i <= 71809) or i == 71900 or (i >= 71950 and i <= 71960) or (i >= 72000 and i <= 72009) or (i >=71500 and i <= 71550) or i == 72500 or (i>=73100 and i <= 73500):
            postcode.append('062')
        elif (i >= 71050 and i <= 71409) or (i >= 71600 and i <= 71759) or (i >= 72120 and i <= 72400) or (i >= 73200 and i <= 73400):
            postcode.append('063')
        elif (i >= 75000  and i <= 75260):
            postcode.append('064')
        elif i == 75460 or i == 76100 or i == 76400 or i == 76450:
            postcode.append('065')
        elif i == 77500 or i == 76200 or i == 76300 or i == 78000 or i == 78100 or i == 78200 or i == 78300:
            postcode.append('066')
        elif (i >= 80000  and i <= 80999):
            postcode.append('067')
        elif (i >= 81100 and i <=  81109):
            postcode.append('068')
        elif (i >= 81200  and i <= 81299):
            postcode.append('069')
        elif i == 81300 or i == 81310 or i == 81550 or i == 81560 or i == 81110 or i == 79000 or i == 79100 or i == 79150 or i == 79200 or i== 79250:
            postcode.append('070')
        
        # -----------71-80--------------------- 
        elif (i >= 81700  and i <= 81799):
            postcode.append('071')
        elif (i >= 81000 and i <= 81099) or (i >= 81400 and i <= 81499):
            postcode.append('072')
        elif (i >= 81600 and i <= 81699) or i == 81800 or (i >= 81900 and i <= 81990):
            postcode.append('073')
        elif(i >= 82000  and i <= 82099) or i == 81500:
            postcode.append('074')
        elif (i >= 83000 and i <=  83999 ) or i ==  86400:
            postcode.append('075')
        elif (i >= 84000  and i  <=  84999):
            postcode.append('076')
        elif (i >= 85000 and i <= 85099) or i == 86500:
            postcode.append('077')
        elif (i >= 86000 and i <= 86399) or (i >= 86600 and i <= 86999) or i == 81850:
            postcode.append('078')
        elif(i >= 25000 and i <= 25989) or (i >= 26000 and i <= 26040) or i == 26060:
            postcode.append('079')
        elif i == 25990 or i == 26050 or (i >= 26070 and i <= 26090) or (i >= 26100 and i <= 26999):
            postcode.append('080')

        # -----------81-90--------------------- 
        elif (i >= 27000 and i <= 27699) or (i >= 49000 and i <= 49099):
            postcode.append('081')
        elif (i >= 28000 and i <= 28899) or (i >= 96000 and i <= 69099):
            postcode.append('082')
        elif (i >= 20000 and i <= 21009) or (i >= 21070 and i <= 21090) or (i>=21100 and i <= 21109) or(i >= 24100 and i >= 24109) or (i >= 24200 and i <= 24209):
            postcode.append('083')
        elif (i >= 21010 and i <= 21030) or i == 21060 or (i >= 21200 and i <= 21309) or i == 21450 or i == 21500 or (i >= 22000 and i <= 22020) or (i >= 22100 and i <= 22120) or (i >= 22200 and i <= 22309) or (i >= 23000 and i <= 23050) or (i >= 23100 and i <= 23400):
            postcode.append('084')
        elif i == 21040 or i == 21400 or (i >= 21600 and i <= 21820) or (i >= 24000 and i <= 24009) or (i >= 24050 and i <= 24060) or i == 24300:
            postcode.append('085')
        elif (i >= 15000 and i <= 15999) or i == 16010 or i == 16150:
            postcode.append('086')
        elif (i >= 16020 and i <= 16090) or i == 16100 or (i >= 16200 and i <= 16899) or (i >= 17000 and i <= 17999) or (i >= 18000 and i <= 18999):
            postcode.append('087')
        elif (i >= 87000 and i <= 87999):
            postcode.append('088')
        elif (i >= 88000 and i <= 88899):
            postcode.append('089')
        elif (i >= 89000 and i <=89999):
            postcode.append('090')

        # -----------91-100--------------------- 
        elif (i >= 90000 and i <= 90999):
            postcode.append('091')
        elif (i >= 91000 and i <= 91099) or (i >=  91200 and i <=91399):
            postcode.append('092')
        elif (i >= 91100 and i <= 91199):
            postcode.append('093')
        elif (i >= 93000 and i <= 93049) or (i >=93100 and i <=93300):
            postcode.append('094')
        elif (i >= 93050 and i <= 93099) or (i >=  93350 and i <= 93450):
            postcode.append('095')
        elif(i>= 94000 and i <= 95999):
            postcode.append('096')
        elif (i >= 96000 and i <=  96999):
            postcode.append('097')
        elif (i >= 97000 and i <= 97999):
            postcode.append('098')
        elif (i >= 98000 and i <= 98999 ):
            postcode.append('099')
    
                                                                               
    # print(postcode)
    time1 = datetime.now()
    start_time=time1.strftime("%H:%M:%S")
    date1 = time1.strftime("%d/%b/%Y")

    Total_accounts=len(kk2)
    file=open(r"E:\\IWOC\\LOG_FILES\\reminder2.txt","w+")
    file.write("INTERCITY  MPC (M) Sdn. Bhd.\n")
    file.write("FIRST REMINDER\n")
    file.write("=============================================================\n")
    file.write("Program ID           : FIRST REMINDER.PDS\n")
    file.write("Paper Type           :\n")
   
    # #--------------------------function for barcode------------------------------

    def barcode(acc_no,count):
        now = datetime.now() 
        date22 = now.strftime("%m%d%y")
        barcode_no = '0902' +  str(date22) +  str(acc_no) 
        # print("date :",date) 
        # print(barcode_no,'barcode_nobarcode_nobarcode_no')
        

        number = barcode_no
        my_code = Code39(number, writer=ImageWriter())
        new_code  = 'b' + str(count)
        my_code.save(r"E:\IWOC\barcode_for_rem2\%s" %(new_code))

        from PIL import Image

        img = Image.open(r"E:\IWOC\barcode_for_rem2\%s.png" %(new_code))
        imgCropped = img.crop(box=(4,10,1080,200))
        imgCropped.save(r"E:\IWOC\barcode_for_rem2\%s.png" %(new_code))
        # imgCropped.show()
        
    # # with open(path.abspath(r"%s\C-%s%s%s_0000%s.pdf"%(kk,d,month,y,p)),'wb')  as append_pdf:
    #             # pdf_merger.write(append_pdf)
    # # -----------------Text File Creation end--------------
 
    count = 1
    c = 1
    pdf = FPDF()
    for i in range(0,Total_accounts):
        barcode(df['ILOM_OLD_SEQUENCE'][i],count)
        now = datetime.now() 
        date44 = now.strftime("%m%d%y")
        barcode_no1 = '0902' +  str(date44) +  str(df['ILOM_OLD_SEQUENCE'][i] )
        # print("date :",date) 
        # print(barcode_no1,'barcode_nobarcode_nobarcode_no')
        pdf.add_page()
        pdf.set_auto_page_break(auto=False)
        pdf.set_line_width(6)
        pdf.set_draw_color(r=0, g=0, b=0)
        pdf.line(x1=5, y1=4, x2=205, y2=4)         #1st horizontal line
        pdf.line(x1=5, y1=292, x2=205, y2=292)     #2nd horizontal line 
        pdf.line(5,6,5,290)                         # 1st vertical line
        pdf.line(205,6,205,290)                     # 2nd vertical line    

        pdf.set_font('Times','B',8.5)
        pdf.cell(45)
        pdf.cell(10,10,txt = 'PERATURAN-PERATURAN PEMBAYARAN PINJAMAN/PEMBIAYAAN')
        pdf.ln(1)
        pdf.cell(66)
        pdf.cell(10,14,txt = 'BANK SIMPANAN NASIONAL (BSN)')

        pdf.ln(2)
        pdf.set_font('Times','',8.5)
        pdf.cell(10,30,txt ='Bayaran PINJAMAN / PEMBIYAAN boleh dilbut secara tunai di mana-mana ATM/CDM/Cawangan BSN , pindahan GIRO / GIRO-I Perbakan internet') 
        pdf.ln(1)
        pdf.cell(10,36, txt = '(www.mybsn.com.my) dan Inter Bank Giro(IBG) dari bank-bank lain yang mengambil bahagian')


        pdf.ln(4)
        pdf.set_font('Times','',8.5)
        pdf.cell(10,44,txt ='Bayaran PINJAMAN / PEMBIYAAN juga boleh dilbut menerusi Arhan Tetap akaun simpanan GIRO/GIRO-I.Bagi membolehkan kami mendebit Akaun') 
        pdf.ln(1)
        pdf.cell(10,50, txt = 'Simpanan BSN tuan/puan, sila pastikan amaun tuan/puan mempunyal baki minima yang ditetapkan oleh BSN dan amaun yang perlu dijelaskan.')


        pdf.ln(4)
        pdf.set_font('Times','',8.5)
        pdf.cell(10,58,txt ='Sekiranya pembayaran melalui cek,draf bank atau arahan juruwang hendaklah dilbut atas nama') 
        pdf.cell(105)
        pdf.set_font('Times','B',8.5)
        pdf.cell(10,58,txt ='BANK SIMPANAN NASIONAL') 
        pdf.cell(33)
        pdf.set_font('Times','',8.5)
        pdf.cell(10,58,txt ='dan diplang') 
        pdf.cell(5)
        pdf.set_font('Times','B',8.5)
        pdf.cell(10,58,txt ='"AKAUN') 
        pdf.ln(1)
        pdf.set_font('Times','B',8.5)
        pdf.cell(10,64, txt = 'PENERIMA SAHAJA"')
        pdf.set_font('Times','',8.5)
        pdf.cell(20)
        pdf.cell(10,64, txt = ',Catatkan nama dan nombot akaun Pinjaman/Pembiayaan tuan/puan di belakang cek berkenaan.Bayaran hendaklah sampai kepada')
        pdf.ln(1)
        pdf.cell(10,70, txt = 'kami sebelum tarikh akhir yang ditetapkan.')


        pdf.ln(4)
        pdf.set_font('Times','',8.5)
        pdf.cell(10,76,txt ='Denda lewat bayar/caj perkhidmatan akan dikenakan jika bayaran yang dibuat oleh tuan/puan diterima selepas tarikh akhir yang ditepkan. Jika') 
        pdf.ln(1)
        pdf.cell(10,82, txt = 'tuan/puan telah menjelaskan amaun berkenaan, sila abaikan pemberitahuan ini.')

        pdf.ln(4)
        pdf.set_font('Times','',8.5)
        pdf.cell(10,88,txt ='Jika tuan/puan memerlukan penjelasan selanjutnya berhubung butir/butir yang dinyatakan di hadapan, sila hubungi kam di talian 03-21623222') 
        pdf.ln(1)
        pdf.cell(10,94, txt = 'atau cawangan di mana tuan/puan membuat permohonan pinjaman/pembiyatan tersebut di talian-talian berikut:-')

        pdf.ln(1)
        pdf.cell(10,104,txt = 'Johor')
        pdf.cell(15)
        pdf.cell(10,104,txt = 'Kedah')
        pdf.cell(15)
        pdf.cell(10,104,txt = 'Kelantan')
        pdf.cell(15)
        pdf.cell(10,104,txt = 'N. Sembilan')
        pdf.cell(15)
        pdf.cell(10,104,txt = 'Melaka')
        pdf.cell(15)
        pdf.cell(10,104,txt = 'Pahang')

        pdf.ln(1)
        pdf.cell(10,108,txt = '07-208-3555')
        pdf.cell(15)
        pdf.cell(10,108,txt = '04-7740 444')
        pdf.cell(15)
        pdf.cell(10,108,txt = '09-745-7070')
        pdf.cell(15)
        pdf.cell(10,108,txt = '06-768-6500')
        pdf.cell(15)
        pdf.cell(10,108,txt = '06-289-5800')
        pdf.cell(15)
        pdf.cell(10,108,txt = '09-565-0565')


        pdf.ln(4)
        pdf.cell(10,110,txt = 'Perak')
        pdf.cell(15)
        pdf.cell(10,110,txt = 'P.Pinanag')
        pdf.cell(15)
        pdf.cell(10,110,txt = 'Sabah')
        pdf.cell(15)
        pdf.cell(10,110,txt = 'Saeawak')
        pdf.cell(15)
        pdf.cell(10,110,txt = 'Terengganu')
        pdf.cell(15)
        pdf.cell(10,110,txt = 'Selangor')

        pdf.ln(1)
        pdf.cell(10,114,txt = '05-245-2222')
        pdf.cell(15)
        pdf.cell(10,114,txt = '04-222-6400')
        pdf.cell(15)
        pdf.cell(10,114,txt = '088-215600')
        pdf.cell(15)
        pdf.cell(10,114,txt = '082-244749')
        pdf.cell(15)
        pdf.cell(10,114,txt = '09-6200 400')
        pdf.cell(15)
        pdf.cell(10,114,txt = '03-5543 3000')

        pdf.ln(1)
        pdf.set_line_width(1)
        pdf.line(10,105,200,105)

        pdf.ln(1)
        pdf.set_font('Arial','',8)
        pdf.cell(90)
        pdf.cell(10,140,txt = 'SULIT')

        pdf.ln(1)
        pdf.cell(10,146, txt = 'JABATAN PEMULIHAN KREDIT')
        pdf.cell(140)
        pdf.cell(10,146,txt = 'TANPA PRASANGKA')

        pdf.ln(1)
        pdf.cell(10,152,txt = 'BANK SIMPANAN NASIONAL')

        pdf.ln(1)
        pdf.cell(10)
        pdf.cell(10,158,txt ="(")
        pdf.cell(-8)
        pdf.cell(10,158,txt =  str(df['br_code'][i]))
        pdf.cell(-3)
        pdf.cell(20,158,txt = ")")


        pdf.ln(1)
        pdf.cell(160)
        pdf.cell(10,164,txt = str(df['SYSDATE_GENERATE'][i]))


        h = 176
        if pd.notna(df['CUSTOMER_NAME'][i]):
            pdf.ln(1)
            pdf.cell(60)
            pdf.cell(10,h,txt = str(df['CUSTOMER_NAME'][i]))

        if pd.notna(df['ADDR1'][i]):
            h = h + 6
            pdf.ln(1)
            pdf.cell(60)
            pdf.cell(10,h,txt = str(df['ADDR1'][i]))

        if pd.notna(df['ADDR2'][i]):
            h = h + 6
            pdf.ln(1)
            pdf.cell(60)
            pdf.cell(10,h,txt = str(df['ADDR2'][i]))

        if pd.notna(df['ADDR3'][i]):
            h = h + 6
            pdf.ln(1)
            pdf.cell(60)
            pdf.cell(10,h,txt = str(df['ADDR3'][i]))

        if pd.notna(df['ADDR4'][i]):
            h = h + 6
            pdf.ln(1)
            pdf.cell(60)
            pdf.cell(10,h,txt = str(df['ADDR4'][i]))

        if pd.notna(df['POSTCODE'][i]):
            h = h + 6
            pdf.ln(1)
            pdf.cell(60)
            pdf.cell(10,h,txt = str(df['POSTCODE'][i]))
            pdf.cell(1)
            pdf.cell(10,h,txt =  str(df['CITY_CD'][i]))

        if pd.notna(df['STATE'][i]):
            h = h + 6
            pdf.ln(1)
            pdf.cell(60)
            pdf.cell(10,h,txt = str(df['STATE'][i]))



        pdf.ln(1)
        pdf.cell(10)
        pdf.cell(10,220, txt = 'Nombor Akaun : ')

        pdf.cell(12)
        pdf.cell(10,220, txt = str(df['ILOM_OLD_SEQUENCE'][i]))

        pdf.cell(80)
        a = str(df['ICBS_PROD_SUBDESC'][i])
        a = a.split()
        a = a[0:5]
        a = " ".join(a)
        pdf.cell(10,220, txt = str(a))

        pdf.ln(1)
        pdf.cell(120)
        b = str(df['ICBS_PROD_SUBDESC'][i])
        b = b.split()
        b= b[5:]
        # print(b)
        b = " ".join(b)
        pdf.cell(10,226, txt = str(b))
        

        pdf.ln(1)
        pdf.cell(10)
        pdf.cell(10,240,txt = 'PERINGATAN 2 ')

        pdf.ln(1)
        pdf.cell(10)
        pdf.cell(10,246,txt = 'NOTIS TUNTUTAN TUNGGAKAN PEMBIAYAAN / PINJAMAN ')


        pdf.ln(1)
        pdf.cell(10)
        pdf.set_font('Arial','',9)
        pdf.cell(10,260,txt = 'Butir - butir pembiayaan / pinjaman Tuan / Puan sehingga tarikh notis tuntutan ini menunjukkan tunggakan sebanyak')


        pdf.ln(1)
        pdf.cell(10)
        pdf.set_font('Arial','',9)
        a = str(df['TOTAL_DUE'][i])
        a = str(a)
        n = len(a)
        if n > 7:
            a = a[0:1] + ',' +  a[1:n] 
            pdf.cell(10,266,txt = 'RM ' + str(a))
        elif n >= 7:
            a = a[0:1] + ',' +  a[1:n] 
            pdf.cell(10,266,txt = 'RM ' + str(a))
        else:
            pdf.cell(10,266,txt = 'RM ' + str(df['TOTAL_DUE'][i]))

        pdf.ln(1)
        pdf.cell(10)
        pdf.set_font('Arial','',9)
        pdf.cell(10,280,txt = 'Sila jelaskan amaun tunggakan dalam tempoh 14 hari dari tarikh notis ini. Hubungi kami melalui talian seperti lampiran')


        pdf.ln(1)
        pdf.cell(10)
        pdf.set_font('Arial','',9)
        pdf.cell(10,286,txt = 'di atas (Seksyen Kutipan) jika Tuan / Puan memerlukan penjelasan lanjut. ')


        pdf.ln(1)
        pdf.cell(10)
        pdf.set_font('Arial','',9)
        pdf.cell(10,300,txt = 'Hubungi Agensi Kaunseling dan Pengurusan Kredit (AKPK) yang menyediakan perkhidmatan pengurusan kewangan,')


        pdf.ln(1)
        pdf.cell(10)
        pdf.set_font('Arial','',9)
        pdf.cell(10,306,txt = 'kaunseling kredit, pendidikan kewangan dan penstrukturan semula pinjaman secara percuma kepada individu. Untuk ')


        pdf.ln(1)
        pdf.cell(10)
        pdf.set_font('Arial','',9)
        pdf.cell(10,312,txt = 'sebarang pertanyaan, sila hubungi talian 1-800-88-2575. ')


        pdf.ln(1)
        pdf.cell(10)
        pdf.set_font('Arial','',9)
        pdf.cell(10,340,txt = 'Ingin bayar bil dan semak baki akaun simpanan anda secara online? Daftarlah Perbankan Internet MyBSN sekarang! ')


        pdf.ln(1)
        pdf.cell(10)
        pdf.set_font('Arial','',9)
        pdf.cell(10,346,txt = 'Kemudahan perbankan hanya untuk anda dari BSN. Untuk maklumat lanjut, layari www.mybsn.com.my atau 1300 88')


        pdf.ln(1)
        pdf.cell(10)
        pdf.set_font('Arial','',9)
        pdf.cell(10,352,txt = '1900. ')

        pdf.ln(1)
        pdf.cell(140)
        pdf.set_font('Arial','',7)
        pdf.cell(10,410,txt = '(1)R2 211122-241122/'  + str(count))


        pdf.add_page()
        pdf.image(r'E:\IWOC\APP SCRIPT\REMINDER2\CYCLE\REMINDER2\DATA\BSN_Reminder-01.jpg',-10,-10,230)

        pdf.set_font('Arial','',7)
        pdf.cell(10,280,txt = '00' + str(c))
        c += 1
        pdf.ln(1)
        pdf.image(r'E:\IWOC\APP SCRIPT\REMINDER2\CYCLE\REMINDER2\DATA\BSN_Reminder-02.jpg', 12,154,180)

        pdf.ln(1)
        new_code  = 'b' + str(count)
        pdf.image(r'E:\IWOC\barcode_for_rem2\%s.png'%(new_code),30,203,70)


        pdf.ln(1)
        pdf.set_font('times','',7)
        pdf.cell(20)
        # pdf.cell(10,412,txt = '090202100670350610301')
        pdf.cell(10,412,txt = str(barcode_no1))
        

        pdf.set_font('times','',7)
        pdf.cell(26)
        pdf.cell(10,412,txt = '00' + str(count))


        pdf.set_font('times','',7)
        pdf.cell(-5)
        pdf.cell(10,412,txt ='(000)')

        
        pdf.set_font('times','',7)
        pdf.cell(-5)
        pdf.cell(10,412,txt =' -P02 - 01')
        
        h = 430
        if pd.notna(df['CUSTOMER_NAME'][i]):
            pdf.ln(1)
            pdf.set_font('Arial','B',10)
            pdf.cell(22)
            pdf.cell(10,h,txt = str(df['CUSTOMER_NAME'][i]))

        if pd.notna(df['ADDR1'][i]):
            h = h + 6 
            pdf.ln(1)
            pdf.cell(22)
            pdf.cell(10,h,txt = str(df['ADDR1'][i]))

        if pd.notna(df['ADDR2'][i]):
            h = h + 6 
            pdf.ln(1)
            pdf.cell(22)
            pdf.cell(10,h,txt = str(df['ADDR2'][i]))

        if pd.notna(df['ADDR3'][i]):
            h = h + 6 
            pdf.ln(1)
            pdf.cell(22)
            pdf.cell(10,h,txt = str(df['ADDR3'][i]))

        if pd.notna(df['ADDR4'][i]):
            h = h + 6 
            pdf.ln(1)
            pdf.cell(22)
            pdf.cell(10,h,txt = str(df['ADDR4'][i]))

        if pd.notna(df['POSTCODE'][i]):
            h = h + 6  
            pdf.ln(1)
            pdf.cell(22)
            pdf.cell(10,h,txt = str(df['POSTCODE'][i]))
            pdf.cell(1)
            pdf.cell(10,h,txt = str(df['CITY_CD'][i]))

        if pd.notna(df['STATE'][i]):
            h = h + 6 
            pdf.ln(1)
            pdf.cell(22)
            pdf.cell(10,h,txt = str(df['STATE'][i]))
        count = int(count)
        count += 1
    
    kk=Folder_creation(cppp)
    bk = kk
    # print(kk,'kkkkkkkkkkkkkkkkkkkkkkkkkkkkk')
        
    pdf.output(r"%s\rem22.pdf"%kk)
    if i == Total_accounts-1:
        rotate_func(bk)
    lesson =(r"%s\rem22.pdf"%kk)
    file_name = os.path.split(lesson)
    ffname = file_name[-1]
    # print(ffname,'saaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaam')
    cppp += 1

    file.write("File Name (PDF)      : ")
    file.write(str(ffname))
    file.write("\n")
    file.write("Data Name (TXT)      : ")
    file.write(str(files1))
    file.write("\n")
    file.write("Total Accounts       :        ")
    file.write(str(Total_accounts))
    file.write("\n")
    file.write("Total Pages          :        ")
    file.write(str(Total_accounts))
    file.write("\n")
    file.write("Total Impression     :        ")
    file.write(str(Total_accounts*2))
    file.write("\n")

    file.write("First Record         : ")
    file.write(str(df['ILOM_OLD_SEQUENCE'][0]))
    file.write("\n")
    file.write("Last Record          : ")
    last_record=df['ILOM_OLD_SEQUENCE'].iat[-1]
    file.write(str(last_record))
    file.write("\n")
    file.write("=============================================================\n")
    file.write("CUSTOMER NAME        : BSN                                   \n")
    file.write("=============================================================\n")
    file.write("Date & Time processed: ")


    file.write(str(date1))
    file.write("    ")
    file.write(str(start_time))
    file.write("    ")


        
    # pdf.output(r'E:\IWOC\APP SCRIPT\REMINDER2\CYCLE\REMINDER2\DATA\rem2.pdf')
    time1 = datetime.now()
    end_time=time1.strftime("%H:%M:%S")
    file.write("    ")
    file.write(str(end_time))
    file.write('\n\n\n\n')
    file.close()


    path20 = r'E:\IWOC\SOURCE' 
    target = r'E:\IWOC\TEMP' 

    for i in move_files:

        src_path = os.path.join(path20, i)
        dst_path = os.path.join(target, i)
        os.rename(src_path, dst_path)
    print('Files Moved................')
